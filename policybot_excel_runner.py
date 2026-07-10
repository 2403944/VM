"""Run multi-turn PolicyBot conversations from an Excel workbook.

Expected workbook columns:
    case_id, turn, question, expected_answer, actual_answer, deepeval, ragas

Rows with the same case_id are treated as one multi-turn conversation. The
script sends each row to POST /query, reusing the returned session_id within
that case_id. It writes:
    - actual_answer: PolicyBot's answer
    - expected_answer: PolicyBot's ground_truth
    - ragas: HumanMessage / AIMessage block on the final row of each case
    - deepeval: Turn(...) block on the final row of each case

Usage:
    .\\.venv\\Scripts\\python.exe excel_policy_requests\\policybot_excel_runner.py --init-sample
    .\\.venv\\Scripts\\python.exe excel_policy_requests\\policybot_excel_runner.py
"""
from __future__ import annotations

import argparse
from collections import defaultdict
from pathlib import Path
from typing import Optional

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font


POLICYBOT_URL = "http://localhost:8001"
DEFAULT_TIMEOUT = 60
DEFAULT_WORKBOOK = Path(__file__).resolve().parent / "policybot_multiturn_output_generation.xlsx"
SHEET_NAME = "Policy Questions"

HEADERS = [
    "case_id",
    "turn",
    "question",
    "expected_answer",
    "actual_answer",
    "deepeval",
    "ragas",
]

SAMPLE_ROWS = [
    {
        "case_id": "MULTI-000",
        "turn": 1,
        "question": "How many casual leaves can an employee take in a year?",
    },
    {
        "case_id": "MULTI-000",
        "turn": 2,
        "question": "Can those leaves be carried forward if they are unused?",
    },
    {
        "case_id": "MULTI-000",
        "turn": 3,
        "question": "What should I remember before applying for this leave?",
    },
    {
        "case_id": "MULTI-001",
        "turn": 1,
        "question": "What types of leave are available to employees?",
    },
    {
        "case_id": "MULTI-001",
        "turn": 2,
        "question": "Which of those are usually paid leaves?",
    },
    {
        "case_id": "MULTI-001",
        "turn": 3,
        "question": "How should an employee apply for one of them?",
    },
    {
        "case_id": "MULTI-001",
        "turn": 4,
        "question": "What happens if the leave is taken without prior approval?",
    },
    {
        "case_id": "MULTI-001",
        "turn": 5,
        "question": "Can you summarize the key rules from this leave discussion?",
    },
    {
        "case_id": "MULTI-002",
        "turn": 1,
        "question": "What is the policy for sick leave?",
    },
    {
        "case_id": "MULTI-002",
        "turn": 2,
        "question": "Do employees need to provide a medical certificate?",
    },
    {
        "case_id": "MULTI-002",
        "turn": 3,
        "question": "When should the manager be informed?",
    },
    {
        "case_id": "MULTI-002",
        "turn": 4,
        "question": "Can sick leave be combined with earned leave?",
    },
    {
        "case_id": "MULTI-002",
        "turn": 5,
        "question": "What happens if the employee has exhausted sick leave?",
    },
    {
        "case_id": "MULTI-002",
        "turn": 6,
        "question": "Are there any approval steps for extending the leave?",
    },
    {
        "case_id": "MULTI-002",
        "turn": 7,
        "question": "What records should HR maintain for this leave?",
    },
    {
        "case_id": "MULTI-002",
        "turn": 8,
        "question": "Summarize the complete sick leave process from this discussion.",
    },
]


def create_sample_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(HEADERS)

    for row in SAMPLE_ROWS:
        ws.append([row.get(header, "") for header in HEADERS])

    widths = {
        "A": 14,
        "B": 8,
        "C": 72,
        "D": 72,
        "E": 72,
        "F": 96,
        "G": 96,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def ask_policybot(
    base_url: str,
    question: str,
    session_id: Optional[str],
    timeout: int,
) -> dict:
    payload = {"question": question}
    if session_id:
        payload["session_id"] = session_id

    response = requests.post(f"{base_url.rstrip('/')}/query", json=payload, timeout=timeout)
    response.raise_for_status()
    return response.json()


def header_indexes(ws) -> dict[str, int]:
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    missing = [header for header in HEADERS if header not in headers]
    if missing:
        raise ValueError(f"Workbook is missing required columns: {', '.join(missing)}")
    return {header: headers.index(header) + 1 for header in HEADERS}


def normalize_case_id(value: object, row_number: int) -> str:
    case_id = str(value or "").strip()
    if not case_id:
        raise ValueError(f"Missing case_id in row {row_number}")
    return case_id


def sort_key(item: dict) -> tuple[int, int]:
    try:
        turn = int(item["turn"])
    except (TypeError, ValueError):
        turn = item["row_number"]
    return turn, item["row_number"]


def quoted(value: str) -> str:
    return str(value or "").replace("\\", "\\\\").replace('"', '\\"')


def build_ragas_template(turns: list[dict]) -> str:
    lines = ["["]
    for turn in turns:
        lines.extend(
            [
                "    HumanMessage(",
                f'        content="{quoted(turn["question"])}"',
                "    ),",
                "    AIMessage(",
                f'        content="{quoted(turn["answer"])}"',
                "    ),",
            ]
        )
    lines.append("]")
    return "\n".join(lines)


def build_deepeval_template(turns: list[dict]) -> str:
    lines = ["["]
    for turn in turns:
        lines.extend(
            [
                '        Turn(role="user",',
                f'             content="{quoted(turn["question"])}"),',
                "",
                '        Turn(role="assistant",',
                f'             content="{quoted(turn["answer"])}"),',
                "",
            ]
        )
    if len(lines) > 1 and lines[-1] == "":
        lines.pop()
    lines.append("]")
    return "\n".join(lines)


def collect_cases(ws, columns: dict[str, int]) -> dict[str, list[dict]]:
    cases: dict[str, list[dict]] = defaultdict(list)
    for row_number in range(2, ws.max_row + 1):
        question = str(ws.cell(row_number, columns["question"]).value or "").strip()
        if not question:
            continue

        case_id = normalize_case_id(
            ws.cell(row_number, columns["case_id"]).value,
            row_number,
        )
        cases[case_id].append(
            {
                "row_number": row_number,
                "case_id": case_id,
                "turn": ws.cell(row_number, columns["turn"]).value,
                "question": question,
                "answer": "",
                "ground_truth": "",
            }
        )

    return {case_id: sorted(rows, key=sort_key) for case_id, rows in cases.items()}


def run_workbook(path: Path, base_url: str, timeout: int) -> None:
    wb = load_workbook(path)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Workbook must contain a sheet named '{SHEET_NAME}'")

    ws = wb[SHEET_NAME]
    columns = header_indexes(ws)
    cases = collect_cases(ws, columns)

    for case_id, rows in cases.items():
        session_id = None
        for row in rows:
            row_number = row["row_number"]
            ws.cell(row_number, columns["deepeval"]).value = ""
            ws.cell(row_number, columns["ragas"]).value = ""

            try:
                data = ask_policybot(
                    base_url,
                    row["question"],
                    session_id=session_id,
                    timeout=timeout,
                )
                session_id = data.get("session_id", session_id)
                row["answer"] = data.get("answer", "")
                row["ground_truth"] = data.get("ground_truth", "")
            except Exception as exc:
                row["answer"] = f"ERROR: {exc}"
                row["ground_truth"] = ""

            ws.cell(row_number, columns["actual_answer"]).value = row["answer"]
            ws.cell(row_number, columns["expected_answer"]).value = row["ground_truth"]

        final_row_number = rows[-1]["row_number"]
        ws.cell(final_row_number, columns["ragas"]).value = build_ragas_template(rows)
        ws.cell(final_row_number, columns["deepeval"]).value = build_deepeval_template(rows)
        print(f"Populated multi-turn testcase: {case_id}")

    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Populate multi-turn PolicyBot Excel conversations.",
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK,
        help=f"Workbook path (default: {DEFAULT_WORKBOOK})",
    )
    parser.add_argument(
        "--url",
        default=POLICYBOT_URL,
        help=f"PolicyBot backend URL (default: {POLICYBOT_URL})",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=DEFAULT_TIMEOUT,
        help=f"Request timeout in seconds (default: {DEFAULT_TIMEOUT})",
    )
    parser.add_argument(
        "--init-sample",
        action="store_true",
        help="Create the sample multi-turn workbook and exit.",
    )
    args = parser.parse_args()

    if args.init_sample:
        create_sample_workbook(args.workbook)
        print(f"Sample workbook created: {args.workbook}")
        return

    if not args.workbook.exists():
        create_sample_workbook(args.workbook)
        print(f"Sample workbook created: {args.workbook}")

    run_workbook(args.workbook, args.url, args.timeout)
    print(f"Workbook populated: {args.workbook}")


if __name__ == "__main__":
    main()
